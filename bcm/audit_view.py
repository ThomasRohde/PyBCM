import ttkbootstrap as ttk
from datetime import datetime
from sqlalchemy import select
from ttkbootstrap.tableview import Tableview
from .models import Capability
import pandas as pd
import openpyxl.styles
from tkinter import filedialog
from .dialogs import create_dialog

class AuditLogViewer(ttk.Toplevel):
    def __init__(self, parent, db_ops):
        """Initialize the audit log viewer window."""
        super().__init__(parent)
        self.db_ops = db_ops
        
        # Store reference to parent's event loop if available
        if hasattr(parent, 'loop'):
            self._loop = parent.loop
        
        # Configure window
        self.withdraw()  # Hide window initially
        self.title("Audit Log Viewer")
        self.geometry("1000x600")  # Made wider to accommodate table
        self.iconbitmap("./bcm/business_capability_model.ico")
        
        # Create toolbar frame
        self.toolbar = ttk.Frame(self)
        self.toolbar.pack(fill="x", padx=10, pady=(10,0))
        
        # Add export button
        self.export_btn = ttk.Button(
            self.toolbar,
            text="Export to Excel",
            command=self.export_to_excel,
            bootstyle="success"
        )
        self.export_btn.pack(side="right")
        
        # Create main frame with table
        self.main_frame = ttk.Frame(self)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Define columns with correct properties
        columns = [
            {"text": "Timestamp", "stretch": False, "width": 150},
            {"text": "Operation", "stretch": False, "width": 100},
            {"text": "Capability", "stretch": True, "width": 200},
            {"text": "Changes", "stretch": True, "width": 400}
        ]
        
        # Configure style for taller rows
        style = ttk.Style()
        style.configure('Treeview', rowheight=60)  # Set row height only
        
        # Create table with scrollbars and additional configuration
        self.table = Tableview(
            self.main_frame,
            coldata=columns,
            searchable=True,
            autofit=False,  # Disable autofit for better column control
            height=20,
            bootstyle="primary"
        )
        
        # Configure the Changes column (index 3) with fixed width
        self.table.view.column(3, stretch=True, width=400, anchor="w")
        
        # Pack the table
        self.table.pack(fill="both", expand=True)
        
        # Add initial loading message
        self.table.insert_row("end", ["Loading audit logs...", "", "", ""])
        
        # Load and display logs
        self.after(100, self.load_logs)
        
        # Center window
        self.position_center()

    def position_center(self):
        """Center the window on the screen."""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        self.deiconify()  # Show window after positioning

    def format_changes(self, old_values: dict, new_values: dict) -> str:
        """Format the changes in a readable way."""
        changes = []
        
        if old_values and new_values:
            # Handle updates - show what changed
            for key in set(old_values.keys()) | set(new_values.keys()):
                old_val = old_values.get(key)
                new_val = new_values.get(key)
                if old_val != new_val:
                    if key == 'parent_id':
                        old_name = self.capability_names.get(old_val, f"Unknown (ID: {old_val})") if old_val else "None"
                        new_name = self.capability_names.get(new_val, f"Unknown (ID: {new_val})") if new_val else "None"
                        changes.append(f"Parent: {old_name} → {new_name}")
                    else:
                        changes.append(f"{key}: {old_val} → {new_val}")
        elif new_values:
            # Handle creation - show new values
            for key, value in new_values.items():
                if key == 'parent_id' and value:
                    parent_name = self.capability_names.get(value, f"Unknown (ID: {value})")
                    changes.append(f"Parent: {parent_name}")
                elif key != 'id':  # Skip ID assignments
                    changes.append(f"{key}: {value}")
        elif old_values:
            # Handle deletion - show what was deleted
            for key, value in old_values.items():
                if key == 'parent_id' and value:
                    parent_name = self.capability_names.get(value, f"Unknown (ID: {value})")
                    changes.append(f"Parent was: {parent_name}")
                else:
                    changes.append(f"{key}: {value}")
                    
        # Join changes with spaces instead of newlines for better table display
        return " | ".join(changes)

    async def get_logs(self):
        """Retrieve logs from database."""
        # First get all capabilities to build name mapping
        self.capability_names = {}
        async with await self.db_ops._get_session() as session:
            stmt = select(Capability.id, Capability.name)
            result = await session.execute(stmt)
            for id, name in result:
                self.capability_names[id] = name
        
        logs = await self.db_ops.export_audit_logs()
        
        # Combine CREATE and ID_ASSIGN operations
        combined_logs = []
        create_log = None
        
        for log in logs:
            if log["operation"] == "CREATE":
                create_log = log
            elif log["operation"] == "ID_ASSIGN" and create_log:
                # Merge ID_ASSIGN info into CREATE log
                if create_log["capability_name"] == log["capability_name"]:
                    create_log["capability_id"] = log["capability_id"]
                    continue
            
            if create_log:
                combined_logs.append(create_log)
                create_log = None
            
            if log["operation"] != "ID_ASSIGN":
                combined_logs.append(log)
                
        return combined_logs

    def load_logs(self):
        """Load and display all audit logs."""
        import asyncio
        
        async def load_async():
            try:
                logs = await self.get_logs()
                
                # Use after() to update GUI from the main thread
                self.after(0, lambda: self._populate_table(logs))
                
            except Exception as e:
                print(f"Error loading logs: {str(e)}")
                error_msg = str(e)  # Capture error message
                self.after(0, lambda: self._show_error(error_msg))

        # Get the event loop
        if hasattr(self, '_loop'):
            loop = self._loop
        else:
            loop = asyncio.get_event_loop()
            self._loop = loop

        # Run the async operation
        asyncio.run_coroutine_threadsafe(load_async(), loop)

    def _populate_table(self, logs):
        """Populate the table with logs (runs in main thread)."""
        try:
            # Sort logs by timestamp (oldest first)
            logs.sort(key=lambda x: datetime.fromisoformat(x["timestamp"]).timestamp(), reverse=False)
            
            # Create list of rows for batch insertion
            rows = []
            for log in logs:
                timestamp = datetime.fromisoformat(log["timestamp"]).strftime("%Y-%m-%d %H:%M:%S")
                operation = log["operation"]
                capability = f"{log['capability_name']}"
                if log["capability_id"]:
                    capability += f" (ID: {log['capability_id']})"
                    
                changes = self.format_changes(log["old_values"], log["new_values"])
                
                # Add row to list with wrapped tag for proper text display
                row = [timestamp, operation, capability, changes]
                rows.append(row)

            # Clear the loading message and existing data
            self.table.delete_rows()  # This removes all rows including the loading message
            
            if rows:
                # Insert all rows at once
                self.table.insert_rows("end", rows)
                # Load the data into view
                self.table.load_table_data()
            else:
                # Show "No logs found" message if there are no logs
                self.table.insert_row("end", ["No audit logs found", "", "", ""])
                self.table.load_table_data()
                
        except Exception as e:
            print(f"Error populating table: {str(e)}")
            self._show_error(str(e))

    def _show_error(self, error_message):
        """Show error in table (runs in main thread)."""
        self.table.delete_rows()
        self.table.insert_row("end", ["Error", "", "", error_message])
        
    def export_to_excel(self):
        """Export the current table view to Excel."""
        try:
            # Get data from tableview and convert to list of dicts
            data = []
            for item in self.table.view.get_children():
                try:
                    values = self.table.view.item(item)['values']
                    if len(values) == 4:  # Ensure we have all columns
                        data.append({
                            "Timestamp": str(values[0]),
                            "Operation": str(values[1]),
                            "Capability": str(values[2]),
                            "Changes": str(values[3])
                        })
                except (IndexError, TypeError, AttributeError) as e:
                    print(f"Error processing row: {e}")
                    continue
            
            if not data:
                create_dialog(
                    self,
                    "Export Failed",
                    "No data to export",
                    ok_only=True
                )
                return
                
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Audit Log Export"
            )
            
            if file_path:
                # Export to Excel with enhanced formatting
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Audit Log', startrow=1)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Audit Log']
                    
                    # Define table style
                    table_style = openpyxl.worksheet.table.TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    
                    # Create table
                    tab = openpyxl.worksheet.table.Table(
                        displayName="AuditLogTable",
                        ref=f"A2:D{len(df) + 2}",
                        tableStyleInfo=table_style
                    )
                    worksheet.add_table(tab)
                    
                    # Add title
                    worksheet['A1'] = 'Business Capability Model - Audit Log'
                    title_cell = worksheet['A1']
                    title_cell.font = openpyxl.styles.Font(size=14, bold=True)
                    worksheet.merge_cells('A1:D1')
                    
                    # Adjust column widths
                    worksheet.column_dimensions['A'].width = 20  # Timestamp
                    worksheet.column_dimensions['B'].width = 15  # Operation
                    worksheet.column_dimensions['C'].width = 30  # Capability
                    worksheet.column_dimensions['D'].width = 50  # Changes
                    
                    # Apply styles to all cells
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(
                                vertical='center',
                                wrap_text=True
                            )
                    
                    # Auto-fit row heights based on content
                    for row in range(1, worksheet.max_row + 1):
                        max_height = 0
                        for cell in worksheet[row]:
                            if cell.value:
                                # Calculate required height based on text content and column width
                                text_lines = str(cell.value).count('\n') + 1
                                # Estimate characters per line based on column width
                                chars_per_line = worksheet.column_dimensions[cell.column_letter].width
                                wrapped_lines = len(str(cell.value)) / chars_per_line
                                total_lines = max(text_lines, wrapped_lines)
                                # Approximate height needed (15 points per line)
                                needed_height = max(15, total_lines * 15)
                                max_height = max(max_height, needed_height)
                        worksheet.row_dimensions[row].height = max_height
                    
                    # Freeze the header row
                    worksheet.freeze_panes = 'A3'
                
                # Show success dialog
                create_dialog(
                    self,
                    "Export Complete",
                    f"Audit log exported successfully to:\n{file_path}",
                    ok_only=True
                )
                
        except Exception as e:
            # Show error dialog
            create_dialog(
                self,
                "Export Failed", 
                f"Failed to export audit log:\n{str(e)}",
                ok_only=True
            )
